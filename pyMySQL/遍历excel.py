import os

from openpyxl import load_workbook


def print_data() -> None:
    # rows: int = sheet.max_row  # 获取行数
    # column: int = sheet.max_column  # 获取列数

    # 从左往右，从上往下
    for rows in range(sheet.max_row):
        print("#########################")
        for column in range(sheet.max_column):
            print(sheet.cell(rows + 1, column + 1).value)


def print_rows():
    # 按行获取值
    print("按行获取值")
    for i in sheet.iter_rows():
        print("#########################")
        for j in i:
            print(j.value)


def print_cols():
    # 按列获取值
    print("按列获取值")
    for i in sheet.iter_cols():
        print("#########################")
        for j in i:
            print(j.value)


if __name__ == "__main__":
    workbook = load_workbook(os.getcwd() + "./test.xlsx")
    sheet = workbook.active

    print_cols()
